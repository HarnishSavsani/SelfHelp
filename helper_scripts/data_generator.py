"""
Retail Synthetic Data Generator
Supports: Gemini API | Ollama (local)
Features:
  - Schema generation with user confirmation loop
  - Chronological batch data generation (no overlap)
  - Streams data to Excel sheet-by-sheet as soon as each table is done
  - Exponential backoff on rate limits (waits 2–5 min range)
  - Robust JSON extraction with multi-strategy fallback
"""

import os
import re
import json
import time
import random
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
#  GLOBAL CONFIG — change these to switch providers/models
# ============================================================

# Provider: "gemini" or "ollama"
PROVIDER = "gemini"

# Gemini settings (used when PROVIDER = "gemini")
GEMINI_API_KEY = "AIzaSyDqxyBOW_gYdzIZy6zNnuwsYa-C8Q_z1Q4"
GEMINI_MODEL   = "gemini-2.5-flash"          # or "gemini-1.5-pro", etc.

# Ollama settings (used when PROVIDER = "ollama")
OLLAMA_BASE_URL = "http://localhost:11434"
OLLAMA_MODEL    = "llama3.1:8b"              # any model pulled via `ollama pull`

# Rows generated per API call (keep ≤ 100 to stay within token limits)
BATCH_SIZE = 50

# Output file
OUTPUT_FILE = "retail_data.xlsx"

# ============================================================
#  LLM CALL LAYER
# ============================================================

def call_llm(prompt: str, max_retries: int = 6) -> str:
    """
    Route to the configured provider with exponential backoff on rate-limit errors.
    Waits between 2–5 minutes when backing off (as requested).
    """
    for attempt in range(max_retries):
        try:
            if PROVIDER == "gemini":
                return _call_gemini(prompt)
            elif PROVIDER == "ollama":
                return _call_ollama(prompt)
            else:
                raise ValueError(f"Unknown PROVIDER: '{PROVIDER}'. Use 'gemini' or 'ollama'.")

        except RateLimitError as e:
            wait = min(120 * (2 ** attempt), 300) + random.uniform(0, 30)
            print(f"⏳ Rate limit hit. Waiting {wait:.0f}s before retry {attempt+1}/{max_retries}...")
            time.sleep(wait)

        except requests.exceptions.RequestException as e:
            print(f"🌐 Network error (attempt {attempt+1}): {e}")
            time.sleep(5)

    raise RuntimeError("❌ All retries exhausted. Check API key / network / model availability.")


class RateLimitError(Exception):
    pass


def _call_gemini(prompt: str) -> str:
    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
    )
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.3}
    }
    resp = requests.post(url, json=payload, timeout=60)

    if resp.status_code == 429:
        raise RateLimitError("Gemini rate limit")
    if not resp.ok:
        raise requests.exceptions.RequestException(
            f"Gemini HTTP {resp.status_code}: {resp.text[:300]}"
        )

    data = resp.json()
    try:
        return data["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError) as e:
        raise ValueError(f"Unexpected Gemini response shape: {data}") from e


def _call_ollama(prompt: str) -> str:
    url = f"{OLLAMA_BASE_URL}/api/generate"
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.3}
    }
    resp = requests.post(url, json=payload, timeout=120)

    if resp.status_code == 429:
        raise RateLimitError("Ollama rate limit")
    if not resp.ok:
        raise requests.exceptions.RequestException(
            f"Ollama HTTP {resp.status_code}: {resp.text[:300]}"
        )

    return resp.json().get("response", "")


# ============================================================
#  JSON EXTRACTION (multi-strategy)
# ============================================================

def extract_json(text: str):
    """
    Try progressively looser strategies to pull valid JSON from LLM output.
    Returns a Python object (list or dict).
    """
    text = text.strip()

    # 1. Direct parse
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 2. Strip markdown fences
    cleaned = re.sub(r"```(?:json)?", "", text).strip().rstrip("`").strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # 3. Extract first JSON array
    m = re.search(r"\[.*?\]", cleaned, re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except json.JSONDecodeError:
            pass

    # 4. Extract first JSON object
    m = re.search(r"\{.*?\}", cleaned, re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except json.JSONDecodeError:
            pass

    # 5. Greedy: find outermost balanced bracket
    for start_char, end_char in [("[", "]"), ("{", "}")]:
        idx = cleaned.find(start_char)
        if idx == -1:
            continue
        depth = 0
        for i, ch in enumerate(cleaned[idx:], start=idx):
            if ch == start_char:
                depth += 1
            elif ch == end_char:
                depth -= 1
            if depth == 0:
                candidate = cleaned[idx: i + 1]
                try:
                    return json.loads(candidate)
                except json.JSONDecodeError:
                    break

    raise ValueError("❌ Could not extract valid JSON from LLM response.")


# ============================================================
#  SCHEMA GENERATION
# ============================================================

SCHEMA_PROMPT = """
You are a data architect. Generate a relational schema for retail analytics.

STRICT RULES:
- Return ONLY valid JSON — no explanation, no markdown, nothing else.
- Schema must reflect the problem statement faithfully.
- Include 4–6 tables. Each table must have 5–10 relevant columns.

OUTPUT FORMAT (exactly):
{{
  "tables": [
    {{
      "table_name": "sales",
      "columns": ["date", "store_id", "product_id", "units_sold", "revenue_inr"]
    }}
  ]
}}

Problem Statement:
{problem}

User Refinements:
{refinements}
"""

def generate_schema(problem: str, refinements: list) -> dict:
    prompt = SCHEMA_PROMPT.format(
        problem=problem,
        refinements="\n".join(refinements) if refinements else "(none)"
    )
    for attempt in range(4):
        raw = call_llm(prompt)
        try:
            parsed = extract_json(raw)
            # Validate structure
            if "tables" in parsed and isinstance(parsed["tables"], list):
                return parsed
            raise ValueError("Missing 'tables' key or wrong type.")
        except Exception as e:
            print(f"⚠️  Schema parse failed (attempt {attempt+1}): {e}")
            time.sleep(2)
    raise RuntimeError("❌ Failed to generate a valid schema after 4 attempts.")


# ============================================================
#  DATA GENERATION — single batch
# ============================================================

DATA_PROMPT = """
Generate realistic synthetic retail data for INDIA.

STRICT RULES:
- Return ONLY a JSON array of {n} objects — no explanation, no markdown.
- Use Indian context:
  - Indian names (Amit, Priya, Rahul, Sneha, …)
  - Indian cities: Pune, Mumbai, Bangalore, Delhi, Chennai, Hyderabad, …
  - Indian states: Maharashtra, Karnataka, Tamil Nadu, …
  - Monetary values in INR (₹)
- Data must be internally consistent and realistic.
- All {n} rows must fall within the date range {start_date} to {end_date}.

Table: {table}
Columns: {columns}
Number of rows: {n}
Date range: {start_date} to {end_date}
"""

def generate_batch(table: str, columns: list, start_date: str, end_date: str, n: int) -> list:
    prompt = DATA_PROMPT.format(
        table=table,
        columns=json.dumps(columns),
        n=n,
        start_date=start_date,
        end_date=end_date
    )
    for attempt in range(4):
        raw = call_llm(prompt)
        try:
            data = extract_json(raw)
            if not isinstance(data, list):
                raise ValueError("Expected a JSON array at the top level.")
            return data
        except Exception as e:
            print(f"⚠️  Data parse failed for '{table}' (attempt {attempt+1}): {e}")
            time.sleep(2)
    raise RuntimeError(f"❌ Failed to generate data for table '{table}' after 4 attempts.")


# ============================================================
#  TABLE-LEVEL GENERATION — chronological batches, no overlap
# ============================================================

def generate_table(table_name: str, columns: list, total_rows: int, start_date: str) -> pd.DataFrame:
    """
    Split total_rows into BATCH_SIZE chunks.
    Each chunk advances the date window chronologically with no day overlap.
    """
    all_rows = []
    window_start = datetime.strptime(start_date, "%Y-%m-%d")
    rows_left = total_rows

    while rows_left > 0:
        batch_n = min(BATCH_SIZE, rows_left)
        # Each batch covers 30 days proportionally (or fewer if last batch is small)
        days_in_window = max(1, int(30 * batch_n / BATCH_SIZE))
        window_end = window_start + timedelta(days=days_in_window - 1)

        start_str = window_start.strftime("%Y-%m-%d")
        end_str   = window_end.strftime("%Y-%m-%d")

        print(f"   ↳ batch {len(all_rows)//BATCH_SIZE + 1}: "
              f"{start_str} → {end_str}  ({batch_n} rows)")

        batch = generate_batch(table_name, columns, start_str, end_str, batch_n)
        all_rows.extend(batch)

        # Next window starts the day AFTER this window ends (no overlap)
        window_start = window_end + timedelta(days=1)
        rows_left -= batch_n

        time.sleep(0.4)   # polite pause between calls

    return pd.DataFrame(all_rows)


# ============================================================
#  EXCEL WRITER — stream sheet-by-sheet
# ============================================================

def init_excel(filepath: str):
    """Create an empty workbook so we can append sheets incrementally."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.save(filepath)
    print(f"📄 Initialised output file: {filepath}")


def append_sheet(filepath: str, sheet_name: str, df: pd.DataFrame):
    """
    Append a DataFrame as a new sheet to an existing workbook.
    Applies basic header formatting.
    """
    wb = load_workbook(filepath)

    # Truncate sheet name to Excel's 31-char limit
    sheet_name = sheet_name[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # Header row
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", start_color="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name).replace("_", " ").title())
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width columns
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_values = [str(col_name)] + [str(v) for v in df.iloc[:, col_idx - 1]]
        max_len = min(max(len(v) for v in col_values), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"   ✅ Sheet '{sheet_name}' written ({len(df)} rows) → {filepath}")


# ============================================================
#  MAIN PIPELINE
# ============================================================

def run_pipeline(problem_statement: str, rows_per_table: int, start_date: str = "2025-09-15"):
    print(f"\n🔧 Provider : {PROVIDER.upper()}")
    print(f"🔧 Model    : {GEMINI_MODEL if PROVIDER == 'gemini' else OLLAMA_MODEL}")
    print(f"🔧 Rows/tbl : {rows_per_table}")
    print(f"🔧 Start dt : {start_date}\n")

    # ---- Schema loop ----
    refinements = []
    schema = None

    while True:
        print("🔍 Generating schema…")
        schema = generate_schema(problem_statement, refinements)
        print("\n📐 Schema:")
        print(json.dumps(schema, indent=2))

        confirm = input("\nProceed with this schema? (y / n): ").strip().lower()
        if confirm == "y":
            break
        feedback = input("✏️  Describe changes (e.g., 'add a promotions table'): ").strip()
        refinements.append(feedback)

    # ---- Initialise Excel file up-front ----
    init_excel(OUTPUT_FILE)

    # ---- Generate data table-by-table, write immediately ----
    print("\n📊 Generating data…\n")
    for tbl in schema["tables"]:
        name    = tbl["table_name"]
        columns = tbl["columns"]
        print(f"🗂️  Table: {name}  (columns: {', '.join(columns)})")

        df = generate_table(name, columns, rows_per_table, start_date)
        append_sheet(OUTPUT_FILE, name, df)
        print()

    print(f"\n🎉 Done! File saved: {OUTPUT_FILE}")


# ============================================================
#  ENTRY POINT
# ============================================================

if __name__ == "__main__":
    PROBLEM_STATEMENT = """
    Build a retail analytics system for an Indian multi-store chain.
    We need several months of transactions for meaningful trend analysis.
    Preprocessing includes data cleaning, normalization, and aggregation by SKU and location.
    Privacy concerns are minimal with anonymized sales data.
    Synthetic data generation is recommended for testing varying demand scenarios.
    Include tables for: stores, products, product categories, sales transactions,
    and inventory levels. Use INR currency throughout.
    NOTE - have store of electronics only
    """

    print("=" * 55)
    print("  Retail Synthetic Data Generator")
    print("=" * 55)

    try:
        rows = int(input("Rows per table (e.g., 100 / 300 / 1000): ").strip())
    except ValueError:
        print("Invalid number. Defaulting to 100.")
        rows = 100

    run_pipeline(PROBLEM_STATEMENT, rows_per_table=rows)